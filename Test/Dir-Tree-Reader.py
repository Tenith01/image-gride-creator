import os
import openpyxl

def add_to_sheet(sheet, path):
    # Get the last row in the sheet
    last_row = sheet.max_row

    # Loop through the files in the folder
    for filename in os.listdir(path):
        # Construct the full file path
        file_path = os.path.join(path, filename)

        # Check if the file is a directory
        if os.path.isdir(file_path):
            # Recursively call the function for the sub-folder
            add_to_sheet(sheet, file_path)
        else:
            # Check if the file has a .png extension
            if file_path.endswith(".png"):
                # Write the file path to the next empty row in the sheet
                sheet.cell(row=last_row + 1, column=1).value = file_path
                last_row += 1

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "DFK Assests"

# Start from the current directory
start_path = os.getcwd()

# Add the file paths to the sheet
add_to_sheet(sheet, start_path)

# Save the workbook
workbook.save("file_paths.xlsx")

